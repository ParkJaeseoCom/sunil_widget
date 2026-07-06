"""출결 위젯: 나이스 16기호 기록 — 완전 로컬(외부 전송 없음), 번호만 저장.

순수부(기호표·기록 헬퍼·파서·Excel 빌더)는 Task 2~4, GUI는 Task 5.
"""

from __future__ import annotations

import calendar
import datetime
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PySide6 import QtCore, QtGui, QtWidgets

from teacher_widgets.core.base_widget import BaseWidget
from teacher_widgets.core.config_store import ConfigStore
from teacher_widgets.core.data_local import (
    load_json_with_backup,
    save_json_with_backup,
)
from teacher_widgets.core.responsive import scale_factor, scaled_font_pt
from teacher_widgets.core.roster import roster_numbers

STATUSES = ["결석", "지각", "조퇴", "결과"]
REASONS = ["질병", "미인정", "기타", "출석인정"]

_SYMBOL_ROWS = {
    "결석": ["♡", "♥", "▲", "△"],
    "지각": ["#", "X", "≠", "◁"],
    "조퇴": ["@", "◎", "∽", "▷"],
    "결과": ["☆", "◇", "=", "▽"],
}
SYMBOLS = {
    (status, REASONS[i]): sym
    for status, row in _SYMBOL_ROWS.items()
    for i, sym in enumerate(row)
}


def symbol_for(status: str, reason: str) -> str:
    return SYMBOLS[(status, reason)]


def set_record(data: dict, date_iso: str, number: int, status: str, reason: str) -> None:
    records = data.setdefault("records", {})
    records.setdefault(date_iso, {})[str(number)] = {
        "status": status, "reason": reason,
    }


def clear_record(data: dict, date_iso: str, number: int) -> None:
    day = data.get("records", {}).get(date_iso)
    if not day:
        return
    day.pop(str(number), None)
    if not day:
        data["records"].pop(date_iso, None)


def get_record(data: dict, date_iso: str, number: int) -> dict | None:
    return data.get("records", {}).get(date_iso, {}).get(str(number))


def month_symbols(data: dict, year: int, month: int) -> dict:
    """해당 월의 {(번호, 일): 기호} 매핑."""
    prefix = f"{year:04d}-{month:02d}-"
    out: dict = {}
    for date_iso, day_records in data.get("records", {}).items():
        if not date_iso.startswith(prefix):
            continue
        day = int(date_iso[8:10])
        for num_str, rec in day_records.items():
            sym = SYMBOLS.get((rec.get("status"), rec.get("reason")))
            if sym:
                out[(int(num_str), day)] = sym
    return out


_DATE_RE = re.compile(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일|(\d{1,2})[/.](\d{1,2})")
_NUMBER_RE = re.compile(r"^(\d{1,3})\s*번")
_CLEAR_WORDS = {"취소", "지움", "지우기"}

_PREFIX_TO_REASON = {"질병": "질병", "미인정": "미인정", "기타": "기타", "인정": "출석인정"}


def _build_keyword_map() -> dict:
    table = {"체험학습": ("결석", "출석인정")}
    for status in STATUSES:
        table[status] = (status, "질병")  # 기본 사유
        for prefix, reason in _PREFIX_TO_REASON.items():
            table[prefix + status] = (status, reason)
    return table


_KEYWORDS = _build_keyword_map()


def parse_command(text: str, today: datetime.date,
                  ref_year: int | None = None) -> dict | None:
    """자연어 출결 명령 파싱(결정적 — AI 아님). 실패 시 None."""
    text = text.strip()
    m_num = _NUMBER_RE.match(text)
    if not m_num:
        return None
    number = int(m_num.group(1))
    rest = text[m_num.end():].strip()

    date_iso = None
    m_date = _DATE_RE.search(rest)
    if m_date:
        month = int(m_date.group(1) or m_date.group(3))
        day = int(m_date.group(2) or m_date.group(4))
        year = ref_year or today.year
        try:
            date_iso = datetime.date(year, month, day).isoformat()
        except ValueError:
            return None
        rest = (rest[: m_date.start()] + rest[m_date.end():]).strip()

    word = rest.replace(" ", "")
    if not word:
        return None
    if word in _CLEAR_WORDS:
        return {"number": number, "date": date_iso, "clear": True}
    if word in _KEYWORDS:
        status, reason = _KEYWORDS[word]
        return {"number": number, "date": date_iso,
                "status": status, "reason": reason}
    return None


_WEEKEND_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")


def months_with_records(data: dict) -> list[str]:
    return sorted({d[:7] for d in data.get("records", {})})


def _legend_text() -> str:
    parts = []
    for status in STATUSES:
        cells = " ".join(f"{reason}{SYMBOLS[(status, reason)]}" for reason in REASONS)
        parts.append(f"{status}: {cells}")
    return "  |  ".join(parts)


def build_attendance_workbook(data: dict, numbers: list[int],
                              months: list[str]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    for ym in months:
        year, month = int(ym[:4]), int(ym[5:7])
        last_day = calendar.monthrange(year, month)[1]
        ws = wb.create_sheet(title=ym)
        ws.cell(row=1, column=1, value=f"{year}년 {month}월 출결")
        ws.cell(row=2, column=1, value=_legend_text())
        ws.cell(row=4, column=1, value="번호")
        for day in range(1, last_day + 1):
            cell = ws.cell(row=4, column=day + 1, value=day)
            if datetime.date(year, month, day).weekday() >= 5:
                cell.fill = _WEEKEND_FILL
        symbols = month_symbols(data, year, month)
        for row_idx, number in enumerate(numbers, start=5):
            ws.cell(row=row_idx, column=1, value=number)
            for day in range(1, last_day + 1):
                sym = symbols.get((number, day))
                if sym:
                    ws.cell(row=row_idx, column=day + 1, value=sym)
    return wb


class AttendanceWidget(BaseWidget):
    BASE_SIZE = (520, 420)

    def __init__(self, store: ConfigStore):
        super().__init__("attendance", store)
        self.data_path = Path(store.path).parent / "attendance.json"
        self._data = load_json_with_backup(self.data_path)
        today = datetime.date.today()
        self._year, self._month = today.year, today.month

        # 상단: ◀ 2026년 7월 ▶
        nav = QtWidgets.QHBoxLayout()
        prev_btn = QtWidgets.QPushButton("◀")
        next_btn = QtWidgets.QPushButton("▶")
        prev_btn.setFixedWidth(28)
        next_btn.setFixedWidth(28)
        prev_btn.clicked.connect(lambda: self.go_month(-1))
        next_btn.clicked.connect(lambda: self.go_month(1))
        self.month_label = QtWidgets.QLabel("", alignment=QtCore.Qt.AlignCenter)
        self.month_label.setStyleSheet("font-weight:700; color:#2b2b2b;")
        nav.addWidget(prev_btn)
        nav.addWidget(self.month_label, stretch=1)
        nav.addWidget(next_btn)
        self.content_layout.addLayout(nav)

        # 그리드
        self.table = QtWidgets.QTableWidget()
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.table.verticalHeader().setDefaultSectionSize(20)
        self.table.horizontalHeader().setDefaultSectionSize(24)
        self.table.cellClicked.connect(self._on_cell_clicked)
        self.content_layout.addWidget(self.table, stretch=1)

        # 자연어 입력줄
        self.command_edit = QtWidgets.QLineEdit()
        self.command_edit.setPlaceholderText("예: 7번 6월 29일 결석 · 5번 체험학습 · 7번 취소")
        self.command_edit.returnPressed.connect(self._on_command_entered)
        self.content_layout.addWidget(self.command_edit)
        self.error_label = QtWidgets.QLabel("")
        self.error_label.setStyleSheet("color:#c0392b;")
        self.content_layout.addWidget(self.error_label)

        self.rebuild_table()

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.Antialiasing)
        painter.setBrush(QtGui.QColor(255, 255, 255, 240))
        painter.setPen(QtCore.Qt.NoPen)
        painter.drawRoundedRect(self.rect(), 16, 16)

    def showEvent(self, event):
        super().showEvent(event)
        self.rebuild_table()  # 숨김 중 학급 구성이 바뀌었을 수 있어 재표시 시 동기화

    # --- 번호/월 ---
    def _numbers(self) -> list[int]:
        boys, girls = self.store.get_roster()
        return roster_numbers(boys, girls)

    def go_month(self, delta: int) -> None:
        month = self._month + delta
        year = self._year
        while month < 1:
            month += 12
            year -= 1
        while month > 12:
            month -= 12
            year += 1
        self._year, self._month = year, month
        self.rebuild_table()

    # --- 그리드 ---
    def rebuild_table(self) -> None:
        numbers = self._numbers()
        last_day = calendar.monthrange(self._year, self._month)[1]
        self.month_label.setText(f"{self._year}년 {self._month}월")
        self.table.clear()
        self.table.setRowCount(len(numbers))
        self.table.setColumnCount(last_day)
        self.table.setHorizontalHeaderLabels([str(d) for d in range(1, last_day + 1)])
        self.table.setVerticalHeaderLabels([str(n) for n in numbers])
        weekend_brush = QtGui.QBrush(QtGui.QColor("#eeeeee"))
        symbols = month_symbols(self._data, self._year, self._month)
        for row, number in enumerate(numbers):
            for day in range(1, last_day + 1):
                item = QtWidgets.QTableWidgetItem(symbols.get((number, day), ""))
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                if datetime.date(self._year, self._month, day).weekday() >= 5:
                    item.setBackground(weekend_brush)
                self.table.setItem(row, day - 1, item)

    def cell_symbol(self, number: int, day: int) -> str:
        numbers = self._numbers()
        if number not in numbers:
            return ""
        row = numbers.index(number)
        item = self.table.item(row, day - 1)
        return item.text() if item else ""

    # --- 기록 적용 ---
    def _save(self) -> None:
        save_json_with_backup(self.data_path, self._data)

    def _refresh_cell(self, number: int, date_iso: str) -> None:
        if not date_iso.startswith(f"{self._year:04d}-{self._month:02d}-"):
            return  # 다른 달 — 현재 그리드 무관
        numbers = self._numbers()
        if number not in numbers:
            return  # roster 축소로 그리드에 없는 번호 — 데이터는 이미 저장됨
        day = int(date_iso[8:10])
        rec = get_record(self._data, date_iso, number)
        text = symbol_for(rec["status"], rec["reason"]) if rec else ""
        row = numbers.index(number)
        item = self.table.item(row, day - 1)
        if item:
            item.setText(text)

    def apply_record(self, number: int, date_iso: str, status: str, reason: str) -> None:
        set_record(self._data, date_iso, number, status, reason)
        self._save()
        self._refresh_cell(number, date_iso)

    def apply_clear(self, number: int, date_iso: str) -> None:
        clear_record(self._data, date_iso, number)
        self._save()
        self._refresh_cell(number, date_iso)

    # --- 자연어 입력 ---
    def _confirm_today(self) -> bool:
        """날짜 생략 시 '오늘로 기록할까요?' 확인. 테스트에서 오버라이드."""
        today = datetime.date.today()
        answer = QtWidgets.QMessageBox.question(
            self, "확인", f"오늘({today.month}/{today.day})로 기록할까요?")
        return answer == QtWidgets.QMessageBox.Yes

    def handle_command_text(self, text: str) -> bool:
        self.error_label.setText("")
        self.error_label.setStyleSheet("color:#c0392b;")
        cmd = parse_command(text, datetime.date.today(), ref_year=self._year)
        if cmd is None:
            self.error_label.setText("이해하지 못했어요 — 예: 7번 6월 29일 결석")
            return False
        if cmd["number"] not in self._numbers():
            self.error_label.setText(f"{cmd['number']}번은 학급 구성에 없습니다")
            return False
        date_iso = cmd["date"]
        if date_iso is None:
            if not self._confirm_today():
                return False
            date_iso = datetime.date.today().isoformat()
        if cmd.get("clear"):
            self.apply_clear(cmd["number"], date_iso)
        else:
            self.apply_record(cmd["number"], date_iso, cmd["status"], cmd["reason"])
        return True

    def _on_command_entered(self) -> None:
        if self.handle_command_text(self.command_edit.text()):
            self.command_edit.clear()

    # --- 셀 퀵메뉴 ---
    def _on_cell_clicked(self, row: int, col: int) -> None:
        numbers = self._numbers()
        if row >= len(numbers):
            return  # roster 축소 후 재구성 전의 낡은 그리드 행
        number = numbers[row]
        date_iso = f"{self._year:04d}-{self._month:02d}-{col + 1:02d}"
        menu = QtWidgets.QMenu(self)
        quick = [("결석 ♡", "결석", "질병"), ("체험학습 △", "결석", "출석인정")]
        for label, status, reason in quick:
            menu.addAction(label, lambda s=status, r=reason:
                           self.apply_record(number, date_iso, s, r))
        menu.addSeparator()
        detail_status = [("결석 상세", "결석"), ("지각", "지각"),
                         ("조퇴", "조퇴"), ("결과", "결과")]
        for label, status in detail_status:
            sub = menu.addMenu(label)
            for reason in REASONS:
                sub.addAction(
                    f"{reason} {symbol_for(status, reason)}",
                    lambda s=status, r=reason:
                    self.apply_record(number, date_iso, s, r))
        menu.addSeparator()
        menu.addAction("지우기", lambda: self.apply_clear(number, date_iso))
        menu.exec(QtGui.QCursor.pos())

    # --- Excel ---
    def export_excel(self, all_months: bool) -> None:
        months = (months_with_records(self._data) if all_months
                  else [f"{self._year:04d}-{self._month:02d}"])
        if not months:
            self.error_label.setText("내보낼 기록이 없습니다")
            return
        default = f"출결_{months[0]}{'_전체' if all_months else ''}.xlsx"
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Excel로 내보내기", default, "Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = build_attendance_workbook(self._data, self._numbers(), months)
            wb.save(path)
            self.error_label.setStyleSheet("color:#2b6e4f;")
            self.error_label.setText(f"저장됨: {path}")
        except OSError as exc:
            self.error_label.setStyleSheet("color:#c0392b;")
            self.error_label.setText(f"저장 실패: {exc}")

    def _custom_menu_actions(self, menu) -> dict:
        cur = menu.addAction("Excel 내보내기(현재 월)")
        full = menu.addAction("Excel 내보내기(전체)")
        return {cur: lambda: self.export_excel(False),
                full: lambda: self.export_excel(True)}

    # --- 반응형 ---
    def _apply_responsive(self) -> None:
        factor = scale_factor((self.width(), self.height()), self.BASE_SIZE)
        self.month_label.setStyleSheet(
            f"font-weight:700; color:#2b2b2b; font-size:{scaled_font_pt(12, factor)}pt;")
        font = self.table.font()
        font.setPointSize(scaled_font_pt(9, factor))
        self.table.setFont(font)

    def on_resized(self, width: int, height: int) -> None:
        self._apply_responsive()
